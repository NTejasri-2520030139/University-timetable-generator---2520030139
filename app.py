"""
app.py — Smart University Timetable Generator
Flask application factory with all routes and blueprints.
"""

import os
import json
import sqlite3
import io
from datetime import datetime
from functools import wraps

from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, jsonify, send_file, g)
from werkzeug.security import generate_password_hash, check_password_hash

# ── App Config ─────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = 'smarttimetable_secret_2024_co3_csp'
DATABASE = os.path.join(os.path.dirname(__file__), 'database.db')

# ── DB Helpers ─────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    db = get_db()
    db.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fullname TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL DEFAULT 'student',
            created_at TEXT DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS subjects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_name TEXT NOT NULL,
            credits INTEGER DEFAULT 3,
            hours_per_week INTEGER DEFAULT 3
        );
        CREATE TABLE IF NOT EXISTS teachers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            teacher_name TEXT NOT NULL,
            specialization TEXT
        );
        CREATE TABLE IF NOT EXISTS classrooms (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            room_number TEXT NOT NULL,
            capacity INTEGER DEFAULT 60
        );
        CREATE TABLE IF NOT EXISTS timeslots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            day TEXT NOT NULL,
            start_time TEXT NOT NULL,
            end_time TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS timetable (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            subject_id INTEGER,
            teacher_id INTEGER,
            classroom_id INTEGER,
            timeslot_id INTEGER,
            FOREIGN KEY(subject_id) REFERENCES subjects(id),
            FOREIGN KEY(teacher_id) REFERENCES teachers(id),
            FOREIGN KEY(classroom_id) REFERENCES classrooms(id),
            FOREIGN KEY(timeslot_id) REFERENCES timeslots(id)
        );
        CREATE TABLE IF NOT EXISTS activity_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            action TEXT NOT NULL,
            detail TEXT,
            timestamp TEXT DEFAULT CURRENT_TIMESTAMP
        );
    ''')
    db.commit()

    # Seed default timeslots if empty
    count = db.execute('SELECT COUNT(*) FROM timeslots').fetchone()[0]
    if count == 0:
        days = ['Monday','Tuesday','Wednesday','Thursday','Friday']
        slots = [('09:00','10:00'),('10:00','11:00'),('11:00','12:00'),
                 ('13:00','14:00'),('14:00','15:00'),('15:00','16:00')]
        for day in days:
            for s,e in slots:
                db.execute('INSERT INTO timeslots (day,start_time,end_time) VALUES (?,?,?)',(day,s,e))
        db.commit()

    db.close()

# ── Auth Decorators ────────────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in first.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def teacher_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'teacher':
            flash('Access denied. Teacher account required.', 'danger')
            return redirect(url_for('student_dashboard'))
        return f(*args, **kwargs)
    return decorated

def student_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def log_activity(action, detail=''):
    if 'user_id' in session:
        db = get_db()
        db.execute('INSERT INTO activity_log (user_id,action,detail,timestamp) VALUES (?,?,?,?)',
                   (session['user_id'], action, detail, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
        db.commit()
        db.close()

# ── Public Routes ──────────────────────────────────────────────────────────

@app.route('/')
def landing():
    return render_template('landing.html')

@app.route('/login', methods=['GET','POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        email = request.form.get('email','').strip().lower()
        password = request.form.get('password','')
        db = get_db()
        user = db.execute('SELECT * FROM users WHERE email=?',(email,)).fetchone()
        db.close()
        if user and check_password_hash(user['password'], password):
            session['user_id'] = user['id']
            session['fullname'] = user['fullname']
            session['email'] = user['email']
            session['role'] = user['role']
            log_activity('LOGIN', f"{user['role']} logged in")
            flash(f"Welcome back, {user['fullname']}!", 'success')
            return redirect(url_for('dashboard'))
        flash('Invalid email or password.', 'danger')
    return render_template('login.html')

@app.route('/register', methods=['GET','POST'])
def register():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        fullname = request.form.get('fullname','').strip()
        email    = request.form.get('email','').strip().lower()
        password = request.form.get('password','')
        confirm  = request.form.get('confirm_password','')
        role     = request.form.get('role','student')

        if not fullname or not email or not password:
            flash('All fields are required.', 'danger')
        elif password != confirm:
            flash('Passwords do not match.', 'danger')
        elif len(password) < 6:
            flash('Password must be at least 6 characters.', 'danger')
        else:
            db = get_db()
            existing = db.execute('SELECT id FROM users WHERE email=?',(email,)).fetchone()
            if existing:
                flash('Email already registered.', 'danger')
                db.close()
            else:
                db.execute('INSERT INTO users (fullname,email,password,role) VALUES (?,?,?,?)',
                           (fullname, email, generate_password_hash(password), role))
                db.commit()
                db.close()
                flash('Account created! Please log in.', 'success')
                return redirect(url_for('login'))
    return render_template('register.html')

@app.route('/logout')
def logout():
    log_activity('LOGOUT')
    session.clear()
    flash('Logged out successfully.', 'info')
    return redirect(url_for('landing'))

@app.route('/dashboard')
@login_required
def dashboard():
    if session.get('role') == 'teacher':
        return redirect(url_for('teacher_dashboard'))
    return redirect(url_for('student_dashboard'))

# ── Teacher Routes ─────────────────────────────────────────────────────────

@app.route('/teacher/dashboard')
@teacher_required
def teacher_dashboard():
    db = get_db()
    stats = {
        'total_subjects':   db.execute('SELECT COUNT(*) FROM subjects').fetchone()[0],
        'total_teachers':   db.execute('SELECT COUNT(*) FROM teachers').fetchone()[0],
        'total_classrooms': db.execute('SELECT COUNT(*) FROM classrooms').fetchone()[0],
        'total_entries':    db.execute('SELECT COUNT(*) FROM timetable').fetchone()[0],
    }

    # Conflict detection
    entries = db.execute('''
        SELECT tt.*, ts.day, ts.start_time, t.teacher_name, c.room_number
        FROM timetable tt
        JOIN timeslots ts ON tt.timeslot_id=ts.id
        JOIN teachers t ON tt.teacher_id=t.id
        JOIN classrooms c ON tt.classroom_id=c.id
    ''').fetchall()
    conflicts = _detect_conflicts([dict(e) for e in entries])
    stats['conflicts'] = len(conflicts)

    # Charts data
    workload = db.execute('''
        SELECT t.teacher_name, COUNT(tt.id) as cnt
        FROM teachers t LEFT JOIN timetable tt ON t.id=tt.teacher_id
        GROUP BY t.id ORDER BY cnt DESC LIMIT 8
    ''').fetchall()
    classroom_util = db.execute('''
        SELECT c.room_number, COUNT(tt.id) as cnt
        FROM classrooms c LEFT JOIN timetable tt ON c.id=tt.classroom_id
        GROUP BY c.id ORDER BY cnt DESC LIMIT 8
    ''').fetchall()
    days = ['Monday','Tuesday','Wednesday','Thursday','Friday']
    weekly = []
    for day in days:
        cnt = db.execute('''SELECT COUNT(*) FROM timetable tt
            JOIN timeslots ts ON tt.timeslot_id=ts.id WHERE ts.day=?''',(day,)).fetchone()[0]
        weekly.append({'day': day, 'count': cnt})

    activities = db.execute('''
        SELECT al.*, u.fullname, u.role FROM activity_log al
        JOIN users u ON al.user_id=u.id
        ORDER BY al.timestamp DESC LIMIT 15
    ''').fetchall()

    db.close()
    return render_template('teacher/dashboard.html',
        stats=stats, conflicts=conflicts,
        workload=[dict(w) for w in workload],
        classroom_util=[dict(c) for c in classroom_util],
        weekly=weekly,
        activities=[dict(a) for a in activities])

# ── Subjects ──

@app.route('/teacher/subjects', methods=['GET','POST'])
@teacher_required
def teacher_subjects():
    db = get_db()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            name  = request.form.get('subject_name','').strip()
            cred  = int(request.form.get('credits', 3))
            hours = int(request.form.get('hours_per_week', 3))
            if name:
                db.execute('INSERT INTO subjects (subject_name,credits,hours_per_week) VALUES (?,?,?)',(name,cred,hours))
                db.commit()
                log_activity('ADD_SUBJECT', name)
                flash(f'Subject "{name}" added.', 'success')
        elif action == 'delete':
            sid = request.form.get('id')
            row = db.execute('SELECT subject_name FROM subjects WHERE id=?',(sid,)).fetchone()
            db.execute('DELETE FROM subjects WHERE id=?',(sid,))
            db.execute('DELETE FROM timetable WHERE subject_id=?',(sid,))
            db.commit()
            log_activity('DELETE_SUBJECT', row['subject_name'] if row else sid)
            flash('Subject deleted.', 'info')
        elif action == 'edit':
            sid   = request.form.get('id')
            name  = request.form.get('subject_name','').strip()
            cred  = int(request.form.get('credits', 3))
            hours = int(request.form.get('hours_per_week', 3))
            db.execute('UPDATE subjects SET subject_name=?,credits=?,hours_per_week=? WHERE id=?',(name,cred,hours,sid))
            db.commit()
            log_activity('EDIT_SUBJECT', name)
            flash('Subject updated.', 'success')
        db.close()
        return redirect(url_for('teacher_subjects'))
    subjects = db.execute('SELECT * FROM subjects ORDER BY subject_name').fetchall()
    db.close()
    return render_template('teacher/subjects.html', subjects=[dict(s) for s in subjects])

# ── Teachers ──

@app.route('/teacher/teachers', methods=['GET','POST'])
@teacher_required
def teacher_teachers():
    db = get_db()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            name = request.form.get('teacher_name','').strip()
            spec = request.form.get('specialization','').strip()
            if name:
                db.execute('INSERT INTO teachers (teacher_name,specialization) VALUES (?,?)',(name,spec))
                db.commit()
                log_activity('ADD_TEACHER', name)
                flash(f'Teacher "{name}" added.', 'success')
        elif action == 'delete':
            tid = request.form.get('id')
            row = db.execute('SELECT teacher_name FROM teachers WHERE id=?',(tid,)).fetchone()
            db.execute('DELETE FROM teachers WHERE id=?',(tid,))
            db.execute('DELETE FROM timetable WHERE teacher_id=?',(tid,))
            db.commit()
            log_activity('DELETE_TEACHER', row['teacher_name'] if row else tid)
            flash('Teacher deleted.', 'info')
        elif action == 'edit':
            tid  = request.form.get('id')
            name = request.form.get('teacher_name','').strip()
            spec = request.form.get('specialization','').strip()
            db.execute('UPDATE teachers SET teacher_name=?,specialization=? WHERE id=?',(name,spec,tid))
            db.commit()
            log_activity('EDIT_TEACHER', name)
            flash('Teacher updated.', 'success')
        db.close()
        return redirect(url_for('teacher_teachers'))
    teachers = db.execute('SELECT * FROM teachers ORDER BY teacher_name').fetchall()
    db.close()
    return render_template('teacher/teachers.html', teachers=[dict(t) for t in teachers])

# ── Classrooms ──

@app.route('/teacher/classrooms', methods=['GET','POST'])
@teacher_required
def teacher_classrooms():
    db = get_db()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'add':
            room = request.form.get('room_number','').strip()
            cap  = int(request.form.get('capacity', 60))
            if room:
                db.execute('INSERT INTO classrooms (room_number,capacity) VALUES (?,?)',(room,cap))
                db.commit()
                log_activity('ADD_CLASSROOM', room)
                flash(f'Room "{room}" added.', 'success')
        elif action == 'delete':
            cid = request.form.get('id')
            row = db.execute('SELECT room_number FROM classrooms WHERE id=?',(cid,)).fetchone()
            db.execute('DELETE FROM classrooms WHERE id=?',(cid,))
            db.execute('DELETE FROM timetable WHERE classroom_id=?',(cid,))
            db.commit()
            log_activity('DELETE_CLASSROOM', row['room_number'] if row else cid)
            flash('Classroom deleted.', 'info')
        elif action == 'edit':
            cid  = request.form.get('id')
            room = request.form.get('room_number','').strip()
            cap  = int(request.form.get('capacity', 60))
            db.execute('UPDATE classrooms SET room_number=?,capacity=? WHERE id=?',(room,cap,cid))
            db.commit()
            log_activity('EDIT_CLASSROOM', room)
            flash('Classroom updated.', 'success')
        db.close()
        return redirect(url_for('teacher_classrooms'))
    classrooms = db.execute('SELECT * FROM classrooms ORDER BY room_number').fetchall()
    db.close()
    return render_template('teacher/classrooms.html', classrooms=[dict(c) for c in classrooms])

# ── Timetable ──

@app.route('/teacher/timetable')
@teacher_required
def teacher_timetable():
    db = get_db()
    timetable = db.execute('''
        SELECT tt.id, s.subject_name, s.credits,
               t.teacher_name, t.specialization,
               c.room_number, c.capacity,
               ts.day, ts.start_time, ts.end_time,
               tt.subject_id, tt.teacher_id, tt.classroom_id, tt.timeslot_id
        FROM timetable tt
        JOIN subjects s ON tt.subject_id=s.id
        JOIN teachers t ON tt.teacher_id=t.id
        JOIN classrooms c ON tt.classroom_id=c.id
        JOIN timeslots ts ON tt.timeslot_id=ts.id
        ORDER BY ts.day, ts.start_time
    ''').fetchall()
    entries = [dict(r) for r in timetable]
    conflicts = _detect_conflicts(entries)

    subjects   = [dict(r) for r in db.execute('SELECT * FROM subjects').fetchall()]
    teachers   = [dict(r) for r in db.execute('SELECT * FROM teachers').fetchall()]
    classrooms = [dict(r) for r in db.execute('SELECT * FROM classrooms').fetchall()]
    timeslots  = [dict(r) for r in db.execute('SELECT * FROM timeslots ORDER BY day,start_time').fetchall()]
    db.close()

    days = ['Monday','Tuesday','Wednesday','Thursday','Friday']
    slots_unique = sorted(set((ts['start_time'],ts['end_time']) for ts in timeslots))
    conflict_ids = set()
    for c in conflicts:
        conflict_ids.update(c.get('entries',[]))

    return render_template('teacher/timetable.html',
        entries=entries, conflicts=conflicts, conflict_ids=conflict_ids,
        subjects=subjects, teachers=teachers, classrooms=classrooms,
        timeslots=timeslots, days=days, slots_unique=slots_unique)

@app.route('/teacher/timetable/generate', methods=['POST'])
@teacher_required
def generate_timetable():
    from algorithms.backtracking import generate_timetable as csp_generate
    db = get_db()
    subjects   = [dict(r) for r in db.execute('SELECT * FROM subjects').fetchall()]
    teachers   = [dict(r) for r in db.execute('SELECT * FROM teachers').fetchall()]
    classrooms = [dict(r) for r in db.execute('SELECT * FROM classrooms').fetchall()]
    timeslots  = [dict(r) for r in db.execute('SELECT * FROM timeslots').fetchall()]

    result = csp_generate(subjects, teachers, classrooms, timeslots)

    if result['success']:
        db.execute('DELETE FROM timetable')
        for a in result['assignments']:
            db.execute(
                'INSERT INTO timetable (subject_id,teacher_id,classroom_id,timeslot_id) VALUES (?,?,?,?)',
                (a['subject_id'], a['teacher_id'], a['classroom_id'], a['timeslot_id'])
            )
        db.commit()
        log_activity('GENERATE_TIMETABLE',
                     f"Solved in {result['metrics'].get('solve_time_ms',0)}ms, "
                     f"{result['metrics'].get('backtracks',0)} backtracks")
        flash(f"✅ Timetable generated successfully! "
              f"Solved in {result['metrics'].get('solve_time_ms',0)}ms with "
              f"{result['metrics'].get('backtracks',0)} backtracks.", 'success')
    else:
        flash(f"❌ {result.get('error','Could not generate timetable.')}", 'danger')

    db.close()
    return jsonify(result)

@app.route('/teacher/timetable/delete/<int:entry_id>', methods=['POST'])
@teacher_required
def delete_timetable_entry(entry_id):
    db = get_db()
    db.execute('DELETE FROM timetable WHERE id=?',(entry_id,))
    db.commit()
    db.close()
    log_activity('DELETE_ENTRY', f'Entry #{entry_id}')
    flash('Entry deleted.', 'info')
    return redirect(url_for('teacher_timetable'))

@app.route('/teacher/timetable/clear', methods=['POST'])
@teacher_required
def clear_timetable():
    db = get_db()
    db.execute('DELETE FROM timetable')
    db.commit()
    db.close()
    log_activity('CLEAR_TIMETABLE')
    flash('Timetable cleared.', 'warning')
    return redirect(url_for('teacher_timetable'))

@app.route('/teacher/timetable/add', methods=['POST'])
@teacher_required
def add_timetable_entry():
    subject_id   = request.form.get('subject_id')
    teacher_id   = request.form.get('teacher_id')
    classroom_id = request.form.get('classroom_id')
    timeslot_id  = request.form.get('timeslot_id')
    db = get_db()
    db.execute('INSERT INTO timetable (subject_id,teacher_id,classroom_id,timeslot_id) VALUES (?,?,?,?)',
               (subject_id, teacher_id, classroom_id, timeslot_id))
    db.commit()
    db.close()
    log_activity('ADD_ENTRY_MANUAL')
    flash('Entry added manually.', 'success')
    return redirect(url_for('teacher_timetable'))

# ── Export ──

@app.route('/teacher/export/pdf')
@teacher_required
def export_pdf():
    from fpdf import FPDF
    db = get_db()
    entries = db.execute('''
        SELECT s.subject_name, t.teacher_name, c.room_number,
               ts.day, ts.start_time, ts.end_time
        FROM timetable tt
        JOIN subjects s ON tt.subject_id=s.id
        JOIN teachers t ON tt.teacher_id=t.id
        JOIN classrooms c ON tt.classroom_id=c.id
        JOIN timeslots ts ON tt.timeslot_id=ts.id
        ORDER BY ts.day, ts.start_time
    ''').fetchall()
    db.close()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 12, 'Smart University Timetable', ln=True, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 8, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', ln=True, align='C')
    pdf.ln(4)

    # Table header
    pdf.set_fill_color(79, 70, 229)
    pdf.set_text_color(255,255,255)
    pdf.set_font('Helvetica', 'B', 9)
    cols = [('Day',28),('Time',32),('Subject',52),('Teacher',40),('Room',28)]
    for label, w in cols:
        pdf.cell(w, 9, label, border=1, fill=True)
    pdf.ln()
    pdf.set_text_color(0,0,0)
    pdf.set_font('Helvetica', '', 8)

    for i, e in enumerate(entries):
        fill = i % 2 == 0
        pdf.set_fill_color(240,240,255) if fill else pdf.set_fill_color(255,255,255)
        pdf.cell(28, 8, e['day'][:3], border=1, fill=fill)
        pdf.cell(32, 8, f"{e['start_time']}-{e['end_time']}", border=1, fill=fill)
        pdf.cell(52, 8, e['subject_name'][:22], border=1, fill=fill)
        pdf.cell(40, 8, e['teacher_name'][:18], border=1, fill=fill)
        pdf.cell(28, 8, e['room_number'][:10], border=1, fill=fill)
        pdf.ln()

    pdf_bytes = pdf.output()
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'timetable_{datetime.now().strftime("%Y%m%d")}.pdf'
    )

@app.route('/teacher/export/excel')
@teacher_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    db = get_db()
    entries = db.execute('''
        SELECT s.subject_name, t.teacher_name, c.room_number,
               ts.day, ts.start_time, ts.end_time
        FROM timetable tt
        JOIN subjects s ON tt.subject_id=s.id
        JOIN teachers t ON tt.teacher_id=t.id
        JOIN classrooms c ON tt.classroom_id=c.id
        JOIN timeslots ts ON tt.timeslot_id=ts.id
        ORDER BY ts.day, ts.start_time
    ''').fetchall()
    db.close()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Timetable'

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Smart University Timetable'
    ws['A1'].font = Font(bold=True, size=14, color='4F46E5')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header row
    headers = ['Day','Start Time','End Time','Subject','Teacher','Room']
    header_fill = PatternFill('solid', fgColor='4F46E5')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row_idx, e in enumerate(entries, 3):
        fill_color = 'F0F0FF' if row_idx % 2 == 0 else 'FFFFFF'
        row_fill = PatternFill('solid', fgColor=fill_color)
        for col, val in enumerate([e['day'],e['start_time'],e['end_time'],
                                   e['subject_name'],e['teacher_name'],e['room_number']], 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal='center')

    # Column widths
    for col, width in zip('ABCDEF', [14,12,12,28,22,14]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'timetable_{datetime.now().strftime("%Y%m%d")}.xlsx')

# ── Student Routes ─────────────────────────────────────────────────────────

@app.route('/student/dashboard')
@login_required
def student_dashboard():
    from datetime import date
    today = date.today().strftime('%A')
    db = get_db()
    today_classes = db.execute('''
        SELECT s.subject_name, t.teacher_name, c.room_number,
               ts.day, ts.start_time, ts.end_time
        FROM timetable tt
        JOIN subjects s ON tt.subject_id=s.id
        JOIN teachers t ON tt.teacher_id=t.id
        JOIN classrooms c ON tt.classroom_id=c.id
        JOIN timeslots ts ON tt.timeslot_id=ts.id
        WHERE ts.day=? ORDER BY ts.start_time
    ''',(today,)).fetchall()
    all_entries = db.execute('''
        SELECT s.subject_name, t.teacher_name, c.room_number,
               ts.day, ts.start_time, ts.end_time
        FROM timetable tt
        JOIN subjects s ON tt.subject_id=s.id
        JOIN teachers t ON tt.teacher_id=t.id
        JOIN classrooms c ON tt.classroom_id=c.id
        JOIN timeslots ts ON tt.timeslot_id=ts.id
        ORDER BY ts.day, ts.start_time
    ''').fetchall()
    db.close()
    days = ['Monday','Tuesday','Wednesday','Thursday','Friday']
    return render_template('student/dashboard.html',
        today=today,
        today_classes=[dict(c) for c in today_classes],
        all_entries=[dict(e) for e in all_entries],
        days=days)

@app.route('/student/timetable')
@login_required
def student_timetable():
    db = get_db()
    entries = db.execute('''
        SELECT s.subject_name, s.credits, t.teacher_name, t.specialization,
               c.room_number, c.capacity,
               ts.day, ts.start_time, ts.end_time
        FROM timetable tt
        JOIN subjects s ON tt.subject_id=s.id
        JOIN teachers t ON tt.teacher_id=t.id
        JOIN classrooms c ON tt.classroom_id=c.id
        JOIN timeslots ts ON tt.timeslot_id=ts.id
        ORDER BY ts.day, ts.start_time
    ''').fetchall()
    teachers   = [dict(r) for r in db.execute('SELECT DISTINCT teacher_name FROM teachers').fetchall()]
    subjects   = [dict(r) for r in db.execute('SELECT DISTINCT subject_name FROM subjects').fetchall()]
    classrooms = [dict(r) for r in db.execute('SELECT DISTINCT room_number FROM classrooms').fetchall()]
    db.close()
    days = ['Monday','Tuesday','Wednesday','Thursday','Friday']
    return render_template('student/timetable.html',
        entries=[dict(e) for e in entries],
        teachers=teachers, subjects=subjects, classrooms=classrooms, days=days)

@app.route('/student/export/pdf')
@login_required
def student_export_pdf():
    return export_pdf()

# ── API Endpoints ──────────────────────────────────────────────────────────

@app.route('/api/conflicts')
@login_required
def api_conflicts():
    db = get_db()
    entries = db.execute('''
        SELECT tt.id, tt.teacher_id, tt.classroom_id, tt.timeslot_id,
               ts.day, ts.start_time, t.teacher_name, c.room_number
        FROM timetable tt
        JOIN timeslots ts ON tt.timeslot_id=ts.id
        JOIN teachers t ON tt.teacher_id=t.id
        JOIN classrooms c ON tt.classroom_id=c.id
    ''').fetchall()
    db.close()
    conflicts = _detect_conflicts([dict(e) for e in entries])
    return jsonify({'conflicts': conflicts, 'count': len(conflicts)})

@app.route('/api/timetable')
@login_required
def api_timetable():
    db = get_db()
    entries = db.execute('''
        SELECT s.subject_name, t.teacher_name, c.room_number,
               ts.day, ts.start_time, ts.end_time
        FROM timetable tt
        JOIN subjects s ON tt.subject_id=s.id
        JOIN teachers t ON tt.teacher_id=t.id
        JOIN classrooms c ON tt.classroom_id=c.id
        JOIN timeslots ts ON tt.timeslot_id=ts.id
        ORDER BY ts.day, ts.start_time
    ''').fetchall()
    db.close()
    return jsonify([dict(e) for e in entries])

# ── Helpers ────────────────────────────────────────────────────────────────

def _detect_conflicts(entries):
    conflicts = []
    for i, e1 in enumerate(entries):
        for e2 in entries[i+1:]:
            if e1['day'] == e2['day'] and e1['start_time'] == e2['start_time']:
                if e1.get('teacher_id') == e2.get('teacher_id') and e1.get('teacher_id'):
                    conflicts.append({
                        'type': 'Teacher Clash',
                        'detail': f"{e1['teacher_name']} double-booked on {e1['day']} at {e1['start_time']}",
                        'entries': [e1.get('id'), e2.get('id')]
                    })
                if e1.get('classroom_id') == e2.get('classroom_id') and e1.get('classroom_id'):
                    conflicts.append({
                        'type': 'Room Clash',
                        'detail': f"Room {e1['room_number']} double-booked on {e1['day']} at {e1['start_time']}",
                        'entries': [e1.get('id'), e2.get('id')]
                    })
    return conflicts

# ── Run ────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    init_db()
    app.run(debug=True, port=5000)
